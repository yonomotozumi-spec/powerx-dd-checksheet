#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PowerX 案件チェックシート 生成スクリプト
（総合評価／DD17項目／許認可39法令／輸送確認／系統接続確認 の5タブ）。
ネットワークはしない。住所・緯度経度・自治体名は呼び出し側(Claude)がGSI等で解決して渡す。
reinfolibの自動判定結果は --values values.json で受け取り、DDの「値・判定」を上書きする（既存3タブのFMTは不変）。
系統接続確認の回答は --grid grid.json（{"items":{"1":{"value":"...","judge":"良|注意|リスク","comment":"..."}}}）で受け取る。
「総合評価」タブは各タブの判定セルをCOUNTIF参照するため、xlsx上で入力を進めると自動で集計が更新される。
"""
import argparse, json, os, urllib.parse, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="FCE4D6")
PERMIT_HEADER_FILL = PatternFill("solid", fgColor="FFF2CC")
HILITE = PatternFill("solid", fgColor="FFF2CC")
TRANS_FILL = PatternFill("solid", fgColor="E2EFDA")
PINK = PatternFill("solid", fgColor="FCE4EC")
LINK_FONT = Font(color="1155CC", underline="single", size=10)
HDR_FONT = Font(bold=True, size=11)
SMALL = Font(size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def gmaps(lat, lon): return f"https://www.google.com/maps?q={lat},{lon}"
def gsi_air(lat, lon): return f"https://maps.gsi.go.jp/#18/{lat}/{lon}/&base=ort&ls=ort&disp=1&vs=c1g1j0h0k0l0u0t0z0r0s0m0f1"
def disaportal(lat, lon): return f"https://disaportal.gsi.go.jp/hazardmap/maps/index.html?ll={lat},{lon}&z=16&base=pale"
FLOOD_LS="flood_l2_kaokutoukai_kagan,0.8%7Cflood_l2_kaokutoukai_hanran,0.8%7Cflood_l2_keizoku,0.8%7Cflood_list,0.8%7Cflood_l1,0.8%7Cflood_list_l2,0.75"
DOSHA_LS="dosha_kiken_nadare,0.8%7Cdosha_keikai_jisuberi,0.8%7Cdosha_keikai_dosekiryu,0.8%7Cdosha_keikai_kyukeisha,0.8"
EKI_LS="ekijouka_zenkoku,0.8"
def haz(lat, lon, ls): return f"https://disaportal.gsi.go.jp/hazardmap/maps/index.html?ll={lat},{lon}&z=16&ls={ls}"
def jshis(lat, lon): return f"https://www.j-shis.bosai.go.jp/map/?lang=jp&ll={lat},{lon}&z=14"
def search(q): return "https://www.google.com/search?q=" + urllib.parse.quote(q)

def dms(dec, is_lat):
    hemi = ("N" if dec >= 0 else "S") if is_lat else ("E" if dec >= 0 else "W")
    dec = abs(dec); d = int(dec); mf = (dec - d) * 60; m = int(mf); s = round((mf - m) * 60, 1)
    return f"{d}°{m}'{s}\"{hemi}"

def dd_items(lat, lon, address, muni, pref):
    m = muni or ""; p = pref or ""
    coord = f"{dms(lat, True)} {dms(lon, False)}（{lat}, {lon}）"
    return [
        ("1","住所",address,gmaps(lat,lon),"リンク","住所→座標は代表点（数百m誤差）。緯度経度があればそれを優先。住居表示と地番は要照合。"),
        ("2","座標",coord,gmaps(lat,lon),"リンク","入力座標。地点の中心がずれていないかマップで確認。"),
        ("3","架線の敷地内縦横断","",gsi_air(lat,lon),"リンク,スクショ","航空写真で敷地内の電柱・架線・縦横断の有無を目視確認。"),
        ("4","ハザード・浸水","",haz(lat,lon,FLOOD_LS),"リンク,スクショ","重ねるハザードマップで「洪水浸水想定区域」をON。"),
        ("5","ハザード・土砂災害","",haz(lat,lon,DOSHA_LS),"リンク,スクショ","重ねるハザードマップで「土砂災害（特別）警戒区域」をON。"),
        ("5-2","土石流危険渓流（参考）","",disaportal(lat,lon),"リンク,スクショ","【参考値・APIなし／自動判定不可】重ねるハザードマップで「土砂災害危険箇所（土石流危険渓流）」レイヤをONにして目視確認。旧調査ベースの参考情報で、法的な区域は項目5の（特別）警戒区域を参照。渓流内・流域に該当する場合は砂防・治水部局へ照会。"),
        ("6","液状化危険度","",haz(lat,lon,EKI_LS),"リンク,スクショ","J-SHIS表層地盤＋各都道府県の液状化マップで確認。"),
        ("7","垂直積雪量","",search(f"{p} {m} 垂直積雪量 建築基準法施行細則"),"リンク,スクショ","自治体の建築基準法施行細則で指定値(cm)を確認。"),
        ("8","市街化区域","",search(f"{m} 都市計画情報 区域区分 市街化区域 マップ"),"リンク,スクショ","自治体の都市計画情報(web GIS)で区域区分を確認。"),
        ("9","市街化調整区域","",search(f"{m} 都市計画情報 市街化調整区域 マップ"),"リンク,スクショ","同上。市街化調整区域内かを確認。"),
        ("10","用途地域","",search(f"{m} 用途地域 都市計画情報 マップ"),"リンク,スクショ","自治体の都市計画情報で用途地域を確認（指定なしの場合あり）。"),
        ("11","農地該当","",gmaps(lat,lon),"リンク,スクショ","地点はGoogleマップ航空写真（座標）で確認。農地区分はeMAFF農地ナビ（住所/地番検索）、種別(1/2/3種)は農業委員会へ照会。"),
        ("12","森林・保安林","",search(f"{p} {m} 森林情報 GIS 保安林 地域森林計画対象民有林"),"リンク,スクショ","国土数値情報A13で森林地域区分(国有林/地域森林計画対象民有林/保安林/保安施設地区)を1次自動判定。保安林の指定範囲・解除可否・作業許可は都道府県森林部局へ照会。"),
        ("13","公園","",search(f"{p} {m} 自然公園 区域 都市公園 区域図"),"リンク,スクショ","自然公園・都市公園の区域内かを確認。"),
        ("14","鳥獣保護区","",search(f"{p} 鳥獣保護区 位置図"),"リンク,スクショ","都道府県の鳥獣保護区位置図で確認。"),
        ("15","埋蔵文化財包蔵地","",search(f"{p} {m} 遺跡地図 埋蔵文化財 包蔵地 GIS"),"リンク,スクショ","都道府県/市町村の遺跡地図GISで周知の埋蔵文化財包蔵地かを確認。"),
        ("16","景観区域","",search(f"{m} 景観計画区域 図"),"リンク,スクショ","自治体の景観計画区域図で対象範囲・高さ規制を確認。"),
        ("17","騒音規制","",search(f"{m} 騒音規制法 地域指定 規制基準"),"リンク,スクショ","騒音規制区域の地域区分と規制基準(dB)を確認。"),
        ("18","振動規制","",search(f"{m} 振動規制法 地域指定 規制基準"),"リンク,スクショ","振動規制区域の地域区分と規制基準(dB)を確認。"),
    ]

PERMITS = [
    ("1","国土利用計画法","土地売買等届出","規模要件と一時金の有無",False),
    ("2","建築基準法","建築確認申請","確認が必要な建築物・工作物か",False),
    ("3","都市計画法","開発行為許可","都計区域外1ha以上の開発か",False),
    ("4","都市計画法","建築等制限（風致地区等）","開発許可区域内外の建築制限",False),
    ("5","県大規模土地利用事前指導要綱","県開発指導要綱手続","県条例の有無・対象事業該当",False),
    ("6","市町村土地開発事業指導要綱","指導要綱手続","市町村条例の有無",False),
    ("7","宅地造成及び特定盛土等規制法","宅地造成等工事許可","規制区域か（2024以降拡大）",False),
    ("8","特定都市河川浸水被害対策法","雨水浸透阻害行為許可","事前相談→許可→届出→完了検査",True),
    ("9","県景観条例","景観計画区域内行為届出","工作物高さ要件・対象か",False),
    ("10","市景観条例","景観計画区域内行為届出","市町村条例の有無",False),
    ("11","河川法","河川区域内占用許可","河川区域・保全区域か",False),
    ("12","砂防法","砂防指定地内開発許可","砂防指定地内か",False),
    ("13","地すべり等防止法","地すべり防止区域内開発許認可","防止区域内か",False),
    ("14","急傾斜地崩壊防止法","急傾斜地崩壊危険区域開発許認可","危険区域内か",False),
    ("15","土砂災害防止法","特別警戒区域特定開発許認可","特別警戒区域内か",False),
    ("16","土砂災害危険箇所","危険箇所区域確認","各危険箇所区域内か",False),
    ("17","道路法①","道路使用/占用許可・工事施工承認","道路使用・占用の有無",True),
    ("18","道路法②","特殊車両通行許可","制限値超の大型車両通行（輸送確認と連動）",True),
    ("19","道路法③","認定道路の確認","道路種類・幅員・拡幅計画",False),
    ("20","法定外公共物","法定外公共物占用許可","法定外道路・水路の有無",False),
    ("21","土壌汚染対策法","土地形質変更届出","3000㎡以上の形質変更か",False),
    ("22","騒音規制法","特定建設作業/特定施設届出","騒音規制区域内か",False),
    ("23","振動規制法","特定建設作業/特定施設届出","振動規制区域内か",False),
    ("24","県生活環境保全条例","区域外騒音指定作業/施設届出","県条例の規制地域内か",True),
    ("25","建設リサイクル法","対象建設工事届出","対象工事・資材か（着工7日前）",True),
    ("26","自然公園法","公園区域内行為規制","公園区域内か",False),
    ("27","県立自然公園条例","県立自然公園区域内行為規制","県条例・区域内か",False),
    ("28","自然環境保全法","保全区域内行為規制","保全区域内か",False),
    ("29","県自然環境保全条例","緑地環境保全/自然記念物届出","県条例・対象要件",False),
    ("30","鳥獣保護法","特別保護地区内開発許可","特別保護区域内か",False),
    ("31","県環境影響評価条例","環境影響評価手続","県条例・対象事業該当",False),
    ("32","市環境影響評価条例","環境影響評価手続","市町村条例の有無",False),
    ("33","森林法","伐採届出/林地開発許可","森林計画対象民有林か・1ha超開発か",False),
    ("34","森林法","保安林伐採許可/形質変更許可","保安林か",False),
    ("35","農地法","農地転用許可","農地種別・土地改良区該当。事前相談→転用許可",True),
    ("36","農振法","農振除外手続","農用地区内か",False),
    ("37","消防法","電気/危険物設備の各届出・許可","蓄電所/発電所の消防条例手続",True),
    ("38","文化財保護法","埋蔵文化財包蔵地土木工事届出","周知の埋蔵文化財包蔵地か",False),
    ("39","工場立地法","特定工場設置届出","特定工場に該当するか",False),
]

TRANSPORT_ITEMS = [
    ("1","出荷元（工場/港）","","輸送元の所在地・最寄IC・港湾。"),
    ("2","輸送ルート（概要）","","出荷元→現地の想定ルート。GoogleマップのルートURLを証跡に。"),
    ("3","車両諸元（積載時）","","全長・全幅・全高・総重量（コンテナ型BESS/PCS）。"),
    ("4","道路幅員（隘路）","","ルート上の最小幅員・すれ違い可否。"),
    ("5","重量制限","","橋梁・道路の重量制限（総重量・軸重）。"),
    ("6","高さ制限","","陸橋・トンネル・標識・架線の高さ制限。"),
    ("7","トンネル","","トンネルの有無・断面・危険物/長大トンネル規制。"),
    ("8","橋梁","","橋梁の重量制限・幅員・通行可否。"),
    ("9","急カーブ・勾配・交差点","","大型トレーラーの転回・内輪差・急勾配の可否。"),
    ("10","踏切","","踏切の有無・通過可否。"),
    ("11","特殊車両通行許可","","制限値超過時の特車通行許可の要否（道路法②と連動）。"),
    ("12","現地進入路・荷下ろし","","敷地進入路の幅員・転回スペース・クレーン設置可否。"),
    ("13","通行規制・時間帯","","通学路・時間帯規制・冬季通行止め等。"),
    ("14","総合判定","","輸送可否（可／要対策／不可）と必要対策。"),
]

# 系統接続確認タブ：No/区分/確認項目/値・入力/判定/証跡/コメント。判定はプルダウン（良/注意/リスク/未確認）。
TSO_CAPACITY_LINKS = {
    "北海道": "https://www.hepco.co.jp/network/connect_access/grid_capacity.html",
    "東北":   "https://nw.tohoku-epco.co.jp/publics/index/20/",
    "東京":   "https://www.tepco.co.jp/electricity-grid/infrastructure/provision-of-information/capacity/index-j.html",
    "中部":   "https://powergrid.chuden.co.jp/network/open/akiyouryou/",
    "北陸":   "https://www.rikuden.co.jp/nw_service/akiyouryou.html",
    "関西":   "https://www.kansai-td.co.jp/connect/open/capacity/",
    "中国":   "https://www.energia.co.jp/nw/service/connect/grid_capacity/",
    "四国":   "https://www.yonden.co.jp/nw/akiyouryou/index.html",
    "九州":   "https://www.kyuden.co.jp/td_network_akiyouryou_index.html",
}

def grid_items(tso):
    cap_url = TSO_CAPACITY_LINKS.get(tso) or search("送配電 空き容量マップ 系統")
    return [
        ("1","系統","公表空き容量（増強要否）",cap_url,"空きあり=良／要確認=注意／空きなし・不明=リスク。マップは申込ベースで埋まるため最新性に注意。"),
        ("2","系統","先行申込案件の状況",search("接続検討 申込状況 公表 "+(tso or "")),"先着優先のため、検討中に容量を先取りされるリスクを確認。"),
        ("3","系統","想定電圧階級",cap_url,"6.6〜33kV=良／66kV=注意／154kV以上=リスク（系統影響評価で回答が長期化）。"),
        ("4","系統","ノンファーム型接続該当",search((tso or "")+" ノンファーム型接続 対象系統"),"該当時は出力制御リスクを事業収支に織り込む（注意）。"),
        ("5","系統","検討区分（簡易/詳細）の見込み","","簡易=良／詳細の可能性=注意／不明=リスク。TSOへ事前ヒアリング推奨。"),
        ("6","スケジュール","系統連系検討の申込状況","","回答受取済み・申込済み=良／未申込=注意／未定=リスク。申込〜回答は通常3〜6ヶ月。"),
        ("7","スケジュール","農地転用・開発許可の要否","","不要=良／申請中=注意／必要・未着手=リスク。許認可No.35/36と連動。"),
        ("8","スケジュール","機器（BESS/PCS）調達見通し","","確保済み=良／交渉中=注意／未定=リスク。"),
        ("9","コスト","接続工事費負担金（万円）","","目安1,000万円以下=良／〜3,000万円=注意／超過=リスク（増強の可能性大）。"),
        ("10","コスト","特定負担の有無","","なし=良／一部あり=注意／あり（系統増強）=リスク。"),
        ("11","コスト","造成工事の必要度","","軽微=良／中程度=注意／大規模=リスク。"),
        ("12","コスト","地権者契約状況","","確定（解除条項あり）=良／交渉中=注意／未着手=リスク。接続不可時の白紙解除特約を必ず確認。"),
    ]

def write_grid(ws, tso, pxno, grid_ov=None):
    ws.sheet_view.showGridLines=False
    CHECK_FILL=PatternFill("solid",fgColor="FFF2CC")
    CHECK_FONT=Font(color="9C6500",size=11)
    AUTO_FILL=PatternFill("solid",fgColor="E1F5EE")
    ws["A1"]=f"系統接続確認 {('PX:'+pxno) if pxno else ''}"; ws["A1"].font=Font(bold=True,size=14)
    ws["A2"]="※判定はプルダウン（良/注意/リスク/未確認）。総合評価タブが自動集計します。"; ws["A2"].font=SMALL
    headers=["No.","区分","確認項目","値・入力","判定","証跡（リンク）","コメント"]
    widths=[6,12,28,24,10,40,44]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for c,h in enumerate(headers,1):
        cell=ws.cell(3,c,h); cell.fill=HEADER_FILL; cell.font=HDR_FONT; cell.alignment=CENTER; cell.border=BORDER
    dv=DataValidation(type="list",formula1='"良,注意,リスク,未確認"',allow_blank=True)
    ws.add_data_validation(dv)
    r=4
    for (no,cat,name,url,comment) in grid_items(tso):
        ws.cell(r,1,no).alignment=CENTER
        ws.cell(r,2,cat).alignment=CENTER
        ws.cell(r,3,name).alignment=WRAP
        vcell=ws.cell(r,4,""); vcell.alignment=WRAP
        jcell=ws.cell(r,5,"未確認"); jcell.alignment=CENTER; jcell.fill=CHECK_FILL; jcell.font=CHECK_FONT
        link=ws.cell(r,6,"確認リンク" if url else "")
        if url: link.hyperlink=url; link.font=LINK_FONT
        link.alignment=WRAP
        ws.cell(r,7,comment).alignment=WRAP
        if grid_ov and no in grid_ov:
            g=grid_ov[no]
            if g.get("value"): vcell.value=g["value"]; vcell.fill=AUTO_FILL
            if g.get("judge") in ("良","注意","リスク"):
                jcell.value=g["judge"]; jcell.font=Font(size=11)
                jcell.fill=AUTO_FILL if g["judge"]=="良" else CHECK_FILL
            if g.get("comment"):
                ws.cell(r,7).value=(comment+" / "+g["comment"]) if comment else g["comment"]
        for c in range(1,8): ws.cell(r,c).border=BORDER
        dv.add(ws.cell(r,5).coordinate)
        ws.row_dimensions[r].height=32; r+=1

def write_summary(ws, args, dd_title, dd_count):
    ws.sheet_view.showGridLines=False
    GOOD_FILL=PatternFill("solid",fgColor="E1F5EE")
    ws.column_dimensions["A"].width=2.5
    for col,w in {"B":22,"C":34,"D":11,"E":13,"F":13,"G":13}.items():
        ws.column_dimensions[col].width=w
    ws["B2"]="総合評価（PX案件チェックシート）"; ws["B2"].font=Font(bold=True,size=14)
    ws["B3"]=f"作成日：{datetime.date.today().isoformat()}"; ws["B3"].font=SMALL
    info=[("PX番号",args.pxno or "―"),("地点",args.address),
          ("座標",f"{args.lat}, {args.lon}"),("送配電エリア",(args.tso+"エリア") if args.tso else "―"),
          ("担当者",args.tanto or "―")]
    r=4
    for k,v in info:
        ws.cell(r,2,k).font=HDR_FONT; ws.cell(r,3,v)
        r+=1
    # 判定・スコア（各タブのCOUNTIF参照。数式なので入力を進めると自動更新される）
    dd_rng=f"'{dd_title}'!$E$3:$E${2+dd_count}"
    pm_rng="'許認可チェックシート'!$G$3:$G$41"
    tr_rng="'輸送確認'!$D$4:$D$17"
    gr_rng="'系統接続確認'!$E$4:$E$15"
    ws["B10"]="総合判定"; ws["B10"].font=Font(bold=True,size=12)
    ws["C10"]='=IF(F14+F16+F17>0,"リスク高・再検討推奨",IF(E14+E17>0,"要確認事項あり",IF(D14+E14=0,"未入力","進行推奨")))'
    ws["C10"].font=Font(bold=True,size=13); ws["C10"].fill=GOOD_FILL
    ws["B11"]="系統接続スコア"; ws["B11"].font=Font(bold=True,size=12)
    ws["C11"]='=IF(D14+E14+F14=0,"―",ROUND((2*D14+E14)/(2*(D14+E14+F14))*100,0)&" / 100")'
    ws["C11"].font=Font(bold=True,size=13)
    # カテゴリ別サマリ表
    headers=["カテゴリ","対象数","良・済","注意・要確認","リスク・要対応","スコア/進捗"]
    for c,h in enumerate(headers,2):
        cell=ws.cell(13,c,h); cell.fill=HEADER_FILL; cell.font=HDR_FONT; cell.alignment=CENTER; cell.border=BORDER
    rows=[
        ("系統接続確認",12,
         f'=COUNTIF({gr_rng},"良")', f'=COUNTIF({gr_rng},"注意")', f'=COUNTIF({gr_rng},"リスク")',
         '=IF(D14+E14+F14=0,"―",ROUND((2*D14+E14)/(2*(D14+E14+F14))*100,0))'),
        ("許認可事前確認（DD）",dd_count,
         f'={dd_count}-COUNTIF({dd_rng},"要確認")', f'=COUNTIF({dd_rng},"要確認")', "―",
         f'=ROUND(D15/{dd_count}*100,0)&"%"'),
        ("許認可チェックシート（39法令）",39,
         f'=COUNTIF({pm_rng},"不要")', "=39-D16-F16", f'=COUNTIF({pm_rng},"要")', "―"),
        ("輸送確認",14,
         f'=COUNTIF({tr_rng},"可")', f'=COUNTIF({tr_rng},"要対策")', f'=COUNTIF({tr_rng},"不可")', "―"),
    ]
    r=14
    for (name,total,good,warn,risk,score) in rows:
        ws.cell(r,2,name).alignment=WRAP
        ws.cell(r,3,total).alignment=CENTER
        ws.cell(r,4,good).alignment=CENTER
        ws.cell(r,5,warn).alignment=CENTER
        ws.cell(r,6,risk).alignment=CENTER
        ws.cell(r,7,score).alignment=CENTER
        for c in range(2,8): ws.cell(r,c).border=BORDER
        ws.row_dimensions[r].height=26; r+=1
    notes=[
        "※このシートは各タブの判定セルを数式（COUNTIF）で参照しており、入力を進めると自動で更新されます。",
        "※総合判定の規則：リスク・要対応が1件でもあれば「リスク高」／注意・要対策があれば「要確認」／それ以外は「進行推奨」。",
        "※系統接続スコア =（良×2＋注意×1）÷（判定済み×2）×100。未確認は分母に含めません。",
        "※許認可チェックシートの「要対応」は G列（対応要否）に「要」と入力された法令数です。",
    ]
    r+=1
    for n in notes:
        ws.cell(r,2,n).font=SMALL; r+=1

def write_dd(ws, items, address, pxno, tanto):
    ws.sheet_view.showGridLines=False
    GRAY_FILL=PatternFill("solid",fgColor="F1EFE8")
    GRAY_FONT=Font(color="7F7F7F",bold=True,size=11)
    AUTO_FILL=PatternFill("solid",fgColor="E1F5EE")   # 緑=自動判定済
    CHECK_FILL=PatternFill("solid",fgColor="FFF2CC")  # 黄=要確認
    CHECK_FONT=Font(color="9C6500",size=11)
    ws.column_dimensions["A"].width=2.5
    ws.column_dimensions["B"].width=2.5
    for col,w in {"C":12,"D":26,"E":34,"F":52,"G":30,"H":14}.items():
        ws.column_dimensions[col].width=w
    # 行1 凡例
    lg=ws["C1"]; lg.value="凡例： 緑=自動判定済 ／ 黄=要確認（人が確認・記入） ／ F列「検索リンク」＝開いて内容を確認"; lg.font=Font(size=10,color="7F7F7F")
    # 行2 見出し（C=グレー、D〜G=オレンジ）
    ws["C2"]="担当者"; ws["C2"].fill=GRAY_FILL; ws["C2"].font=GRAY_FONT; ws["C2"].alignment=CENTER; ws["C2"].border=BORDER
    ws["D2"]="地点名称"; ws["D2"].fill=HEADER_FILL; ws["D2"].font=HDR_FONT; ws["D2"].alignment=Alignment(horizontal="left",vertical="center"); ws["D2"].border=BORDER
    ws["E2"]="値・判定"; ws["E2"].fill=HEADER_FILL; ws["E2"].font=GRAY_FONT; ws["E2"].alignment=CENTER; ws["E2"].border=BORDER
    ws["F2"]="証跡"; ws["F2"].fill=HEADER_FILL; ws["F2"].font=GRAY_FONT; ws["F2"].alignment=CENTER; ws["F2"].border=BORDER
    ws["G2"]="コメント"; ws["G2"].fill=HEADER_FILL; ws["G2"].font=GRAY_FONT; ws["G2"].alignment=CENTER; ws["G2"].border=BORDER
    ws["H2"]="確認方法"; ws["H2"].fill=HEADER_FILL; ws["H2"].font=GRAY_FONT; ws["H2"].alignment=CENTER; ws["H2"].border=BORDER
    r=3
    for it in items:
        no,name,val,url,how,comment = it[0],it[1],it[2],it[3],it[4],it[5]
        kind = "search" if "google.com/search" in (url or "") else "map"
        c=ws.cell(r,3,tanto); c.font=Font(color="7F7F7F",size=11); c.alignment=CENTER; c.fill=GRAY_FILL
        ws.cell(r,4,f"{no}. {name}").alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ecell=ws.cell(r,5)
        if val:
            ecell.value=val; ecell.fill=AUTO_FILL
        else:
            ecell.value="要確認"; ecell.fill=CHECK_FILL; ecell.font=CHECK_FONT
        ecell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        if kind=="search":
            disp="検索リンク（参考）" if val else "検索リンク（要確認）"
        else:
            disp=url or ""
        link=ws.cell(r,6,disp)
        if url:
            link.hyperlink=url; link.font=LINK_FONT
        link.alignment=WRAP
        ws.cell(r,7,comment).alignment=WRAP
        ws.cell(r,8,how).alignment=CENTER
        for cc in range(3,9): ws.cell(r,cc).border=BORDER
        ws.row_dimensions[r].height=34
        r+=1

def write_permits(ws, pxno, permit_ov=None):
    ws.sheet_view.showGridLines=False
    ws["A1"]=f"許認可チェックシート {('PX:'+pxno) if pxno else ''}"; ws["A1"].font=Font(bold=True,size=14)
    headers=["重要","担当","No.","法令","手続名","確認の要点","対応要否","対象範囲","確認日","机上","連絡","行政窓口","担当者","連絡先","工程・日数","備考"]
    widths=[6,8,5,24,28,40,9,12,11,7,7,16,12,14,14,20]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for c,h in enumerate(headers,1):
        cell=ws.cell(2,c,h); cell.fill=PERMIT_HEADER_FILL; cell.font=HDR_FONT; cell.alignment=CENTER; cell.border=BORDER
    r=3
    for (no,law,proc,point,important) in PERMITS:
        ws.cell(r,1,"★" if important else "").alignment=CENTER
        ws.cell(r,3,no).alignment=CENTER
        ws.cell(r,4,law).alignment=WRAP; ws.cell(r,5,proc).alignment=WRAP; ws.cell(r,6,point).alignment=WRAP
        ws.cell(r,7,"").alignment=CENTER
        for c in range(8,17): ws.cell(r,c,"")
        if important:
            for c in range(1,17): ws.cell(r,c).fill=HILITE
        if permit_ov and no in permit_ov:
            ws.cell(r,7,permit_ov[no].get("req","要")); ws.cell(r,7).alignment=CENTER
            ws.cell(r,16,permit_ov[no].get("note",""))
            for c in range(1,17): ws.cell(r,c).fill=PINK
        for c in range(1,17): ws.cell(r,c).border=BORDER
        ws.row_dimensions[r].height=28; r+=1
    n=ws.cell(r+1,4,"※「要」の手続は工程（1M～6M＋着工・完工）をガント化。代表例：特定都市河川/農地法/消防法/道路法①②/生活環境保全条例/建設リサイクル法。"); n.font=SMALL

def write_transport(ws, items, address, pxno):
    ws.sheet_view.showGridLines=False
    ws["A1"]=f"輸送確認 {('PX:'+pxno) if pxno else ''}"; ws["A1"].font=Font(bold=True,size=14)
    ws["C1"]=f"納入先：{address}"; ws["C1"].font=SMALL
    ws["A2"]="※コンテナ型蓄電池(BESS)/PCSの輸送可否確認。社内リファレンスに合わせて項目を調整してください。"; ws["A2"].font=SMALL
    headers=["No.","確認項目","確認結果","判定","証跡（リンク）","備考"]
    widths=[6,26,40,10,40,30]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for c,h in enumerate(headers,1):
        cell=ws.cell(3,c,h); cell.fill=TRANS_FILL; cell.font=HDR_FONT; cell.alignment=CENTER; cell.border=BORDER
    r=4
    for (no,name,result,comment) in items:
        ws.cell(r,1,no).alignment=CENTER; ws.cell(r,2,name).alignment=WRAP; ws.cell(r,3,result).alignment=WRAP
        ws.cell(r,4,"").alignment=CENTER; ws.cell(r,5,"").alignment=WRAP; ws.cell(r,6,comment).alignment=WRAP
        for c in range(1,7): ws.cell(r,c).border=BORDER
        ws.row_dimensions[r].height=30; r+=1

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--address",required=True); ap.add_argument("--lat",type=float,required=True); ap.add_argument("--lon",type=float,required=True)
    ap.add_argument("--muni",default=""); ap.add_argument("--pref",default=""); ap.add_argument("--pxno",default=""); ap.add_argument("--tanto",default="")
    ap.add_argument("--values",default=""); ap.add_argument("--out",required=True)
    ap.add_argument("--tso",default="",help="送配電エリア（北海道/東北/東京/中部/北陸/関西/中国/四国/九州）。空き容量マップのリンク生成に使用")
    ap.add_argument("--grid",default="",help="系統接続確認の回答JSON（任意）")
    ap.add_argument("--classic",action="store_true",help="元の3タブ（DD/許認可39/輸送）のみ生成。総合評価・系統接続確認タブを付けない")
    args=ap.parse_args()
    run(args)
    print("OK:",args.out)


def run(args):
    """argparse名前空間(または同等の属性を持つオブジェクト)を受け取りxlsxを生成する。
    app.py からプロセス内で直接呼べるよう main() から分離した。"""
    items=dd_items(args.lat,args.lon,args.address,args.muni,args.pref)
    permit_ov={}
    if args.values and os.path.exists(args.values):
        _data=json.load(open(args.values,encoding="utf-8"))
        ov=_data.get("values",{})
        permit_ov=_data.get("permits",{})
        merged=[]
        for (no,name,val,url,how,comment) in items:
            if no in ov:
                nv=(ov[no].get("value") or "").strip(); nc=(ov[no].get("comment") or "").strip()
                if nv: val=nv
                if nc: comment=(comment+" / "+nc) if comment else nc
            merged.append((no,name,val,url,how,comment))
        items=merged
    wb=Workbook(); ws1=wb.active
    safe=args.address
    for ch in '[]:*?/\\': safe=safe.replace(ch,"")
    ws1.title=(safe[:28]+"…") if len(safe)>28 else (safe or "DD")
    write_dd(ws1,items,args.address,args.pxno,args.tanto)
    write_permits(wb.create_sheet("許認可チェックシート"),args.pxno,permit_ov)
    write_transport(wb.create_sheet("輸送確認"),TRANSPORT_ITEMS,args.address,args.pxno)
    if not args.classic:
        grid_ov={}
        if args.grid and os.path.exists(args.grid):
            grid_ov=json.load(open(args.grid,encoding="utf-8")).get("items",{})
        write_grid(wb.create_sheet("系統接続確認"),args.tso,args.pxno,grid_ov)
        write_summary(wb.create_sheet("総合評価",0),args,ws1.title,len(items))
    wb.save(args.out)
    return args.out

if __name__=="__main__":
    main()
